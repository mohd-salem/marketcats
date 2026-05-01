"""
Excel export service.

Reads the original uploaded file (CSV or Excel), appends relevance and
categorization columns, and returns the workbook as bytes.
AI-generated columns are highlighted with a light-blue header.
"""

import io
from pathlib import Path
from typing import Any

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

AI_HEADER_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
AI_HEADER_FONT = Font(bold=True, color="1F3864")


def _load_original_df(file_path: Path) -> pd.DataFrame:
    suffix = file_path.suffix.lower()
    if suffix in (".xlsx", ".xls"):
        return pd.read_excel(file_path, dtype=str)
    try:
        return pd.read_csv(file_path, dtype=str, encoding="utf-8-sig")
    except UnicodeDecodeError:
        return pd.read_csv(file_path, dtype=str, encoding="latin-1")


def build_excel_export(
    file_path: Path,
    products: list[dict[str, Any]],
    dimensions: list[dict[str, Any]],  # [{"id": int, "name": str}, ...]
) -> bytes:
    """
    Preserve the original file, then append:
      AI Relevance | Relevance Confidence | Relevance Reasoning | Keep/Exclude
      [for each dimension: Dim Name | Dim Value | Dim Confidence]
      Final Notes | Manual Override
    """
    original_df = _load_original_df(file_path)

    by_index: dict[int, dict[str, Any]] = {p["row_index"]: p for p in products}

    # Build relevance columns
    ai_relevance, rel_confidence, rel_reasoning, keep_col = [], [], [], []
    dim_cols: dict[str, list[str]] = {}  # "Name", "Value", "Confidence" per dim
    final_notes_col, manual_override_col = [], []

    for i in range(len(original_df)):
        p = by_index.get(i, {})
        ai_relevance.append(p.get("ai_relevance") or "")
        rel_confidence.append(
            f"{p['ai_relevance_confidence']:.0%}" if p.get("ai_relevance_confidence") is not None else ""
        )
        rel_reasoning.append(p.get("ai_relevance_reasoning") or "")
        keep = p.get("keep")
        keep_col.append("Keep" if keep is True else "Exclude" if keep is False else "")

        for dim in dimensions:
            name = dim["name"]
            cats = p.get("final_categories") or {}
            ai_cats: list[dict] = p.get("ai_categories") or []
            dim_assignment = next((a for a in ai_cats if a.get("dimension_name") == name), {})

            dim_cols.setdefault(f"_name_{name}", []).append(name if cats.get(name) else "")
            dim_cols.setdefault(f"_value_{name}", []).append(cats.get(name, ""))
            dim_cols.setdefault(f"_conf_{name}", []).append(
                f"{dim_assignment['confidence']:.0%}" if dim_assignment.get("confidence") is not None else ""
            )

        final_notes_col.append(p.get("final_notes") or "")
        manual_override_col.append(str(p.get("manual_override") or ""))

    original_df["AI Relevance"] = ai_relevance
    original_df["Relevance Confidence"] = rel_confidence
    original_df["Relevance Reasoning"] = rel_reasoning
    original_df["Keep/Exclude"] = keep_col

    original_col_count = len(original_df.columns)

    for i, dim in enumerate(dimensions, start=1):
        name = dim["name"]
        original_df[f"Category Dimension {i}"] = dim_cols[f"_name_{name}"]
        original_df[f"Category Dimension {i} Value"] = dim_cols[f"_value_{name}"]
        original_df[f"Category Dimension {i} Confidence"] = dim_cols[f"_conf_{name}"]

    original_df["Final Notes"] = final_notes_col
    original_df["Manual Override"] = manual_override_col

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Categorized Products"

    headers = list(original_df.columns)
    ws.append(headers)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        if col_idx > original_col_count:
            cell.fill = AI_HEADER_FILL
            cell.font = AI_HEADER_FONT

    # Write data rows
    for _, row in original_df.iterrows():
        ws.append(list(row))

    # Auto-fit column widths (capped at 50)
    from openpyxl.utils import get_column_letter
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells),
            default=10,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    # Freeze top row
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
